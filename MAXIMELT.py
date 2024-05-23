import numpy as np
import pandas as pd
import os
import tkinter
from tkinter import filedialog
tkinter.Tk().withdraw() # prevents an empty tkinter window from appearing
import time

print("\nMAXIMELT screw generator V1\n")
print("Select Excel file to compute OUTPUT")
current_dir = os.getcwd()
filename = filedialog.askopenfilename(initialdir = current_dir, title = "Select a File", filetypes = (("Excel files", "*.xlsm*"),("all files", "*.*")))

# INPUT % Read inputs from excel file for each zone
df = pd.read_excel(filename, sheet_name='Ric-Input')
NZ = df.iloc[1, (10)]          # Number of zones

# Input lists
CD = df.iloc[3, (3)]    # [-] Codice vite
DA = df.iloc[8, (2)]    # [kg/dm3] Densità apparente alimentazione (pellet) ZONE INDEPENDENT
VR = df.iloc[4, (2)]    # [RPM] Velocità rotazione ZONE INDEPENDENT
RC = df.iloc[7, (2)]    # [-] Rapporto di compressione ZONE INDEPENDENT
DE = df.iloc[9, (2)]    # [kg/dm3] Densità PVC ZONE INDEPENDENT
#NF = df.iloc[25,(2)]   # [-] Numero fresature
#DF = df.iloc[6, (2)]   # [mm] Diametro fresa
LS = df.iloc[3, (2)]    # [mm] Lunghezza vite
DV = list()             # [mm] Diametro vite
PA = list()             # [mm] Passo
NP = list()             # [-] Numero principi
IA = list()             # [mm] Interasse
LZ = list()             # [mm] Lunghezza zona
PF = list()             # [mm] Profondità filetto
CR = list()             # [mm] Lunghezza cresta
FD = list()             # [mm] Lunghezza fondo
LG = list()             # [mm] Length of gap
ID = list()             # [mm] Core diameter

# Additional parameters need to be asked in the input excel file 
DV_1 = df.iloc[5, (10)]                 # Outer diameter input side (larger side) ZONE INDEPENDENT
DV_2 = df.iloc[6, (10)]                 # Outer diameter output side (smaller side) ZONE INDEPENDENT
ID_1 = df.iloc[3, (10)]                 # Core diameter input side ZONE INDEPENDENT
ID_2 = df.iloc[4, (10)]                 # Core diameter output side ZONE INDEPENDENT
IA_1 = df.iloc[7, (10)]                 # Axial distance input side ZONE INDEPENDENT
IA_2 = df.iloc[8, (10)]                 # Axial distance output side ZONE INDEPENDENT   
m_DV = ((DV_2 / 2) - (DV_1 / 2)) / LS   # [-] External diameter slope
m_IA = ((IA_2 / 2) - (IA_1 / 2)) / LS   # [-] Axis distance slope
m_ID = ((ID_2 / 2) - (ID_1 / 2)) / LS   # [-] Inner diameter slope

# Temporary memory
DV_x=list()     
ID_x=list()
PF_x=list()
IA_x=list()

pos=0                   #start position
for i in range(NZ):     # zone changing 
    DV_x = []   
    ID_x = []   
    PF_x = []   
    IA_x = []   

    # zone dependent input parameters
    PA.append(df.iloc[12, (2+i)])                   # [mm] Passo
    NP.append(df.iloc[13, (2+i)])                   # [-] Numero principi
    LZ.append(df.iloc[14, (2+i)])                   # [mm] Lunghezza zona
    CR.append(df.iloc[16, (2+i)])                   # [mm] Lunghezza cresta
    FD.append(df.iloc[17, (2+i)])                   # [mm] Lunghezza fondo
    LG.append(df.iloc[18, (2+i)])                   # [mm] Lunghezza gaps  

    start = pos
    stop =  pos + LZ[i]       
    step = 1
    num_samples = int((stop - start) / step)                      

    for j in np.linspace(start, stop, num_samples):       # Calculation of x dependent parameters
        DV_x.append(2*m_DV * j + DV_1 )             # outer diameter
        ID_x.append(2*m_ID * j + ID_1 )             # core diameter
        PF_x.append((DV_x[-1] - ID_x[-1]) / 2)      # thread depth
        IA_x.append(2*m_IA * j + IA_1)              # axial distance. Control again! we are referring to the system

    DV.append(DV_x)                                 # add DV values with changing x as a sublist. each sublist is a zone.
    PF.append(PF_x)                                 # add PF values with changing x as a sublist. each sublist is a zone.
    IA.append(IA_x)                                 # add IA values with changing x as a sublist. each sublist is a zone.
    ID.append(ID_x)                                 # add ID values with changing x as a sublist. each sublist is a zone.
    
    pos += LZ[i] + (LG[i])

# OUTPUT LIST
CL = list()    # Cresta longitudinale  
FL = list()    # Fondo longitudinale     
LC = list()    # Luce di calandra 
AF = list()    # Angolo di fresa
TL = list()    # Traferro longitudinale
PFc = list()   # Profondità filetto
CX = list()    # Rapporto di compenetrazione X
CO = list()    # Angolo compenetrazione
CP = list()    # Compenetrazione viti %
AC = list()    # Area cilindro
cCR = list()   # Calcolo cresta
cFD = list()   # Calcolo fondo
CT = list()    # Cresta Trasversale
FT = list()    # Fondo Trasversale
AV = list()    # Area vite
EV = list()    # Volume vuoto al giro
LF = list()    # Luce fianchi
LT = list()    # Luce tetraedrica
QM = list()    # Portata massima
TP = list()    # Tempo di permanenza zone piene
TNP = list()   # Tempo di permanenza zone NON piene
PL = list()    # Peso calandrabile
PL2 = list()   # Peso calandrabile2
PC = list()    # Peso di una camera a C
GT = list()    # Massimo gradiente velocità di taglio
CC = list()    # Superficie di contatto cilindro in una zona
CV = list()    # Superficie di contatto con le viti in una zona
CV2 = list()   # Superficie di contatto con le viti in una zona (versione 2)
AEb = list()   # Angolo elica base vite
AEm = list()   # Angolo elica metà vite
AEc = list()   # Angolo elica cresta vite
CPc = list()   # Canale perpendicolare cresta vite
CPb = list()   # Canale perpendicolare base vite
CPm = list()   # Canale perpendicolare metà vite
CRP = list()   # Cresta perpendicolare
CS = list()    # Calandrate statistiche
SR = list()    # Shear rate
Pth = list()   # Potenza teorica del motore
Ath = list()   # Amperes teorici del motore
nC = list()    # Numero creste
Lper = list()  # Lunghezza percentuale
ASR = list()   # Area sezione retta cilindro


for k in range(NZ): #zone counter
    CL_x = []   # Cresta longitudinale  
    FL_x = []   # Fondo longitudinale     
    LC_x = []   # Luce di calandra 
    AF_x = []   # Angolo di fresa
    TL_x = []   # Traferro longitudinale
    PFc_x = []  # Profondità filetto
    CX_x = []   # Rapporto di compenetrazione X
    CO_x = []   # Angolo compenetrazione
    CP_x = []   # Compenetrazione viti %
    AC_x = []   # Area cilindro
    cCR_x = []  # Calcolo cresta
    cFD_x = []  # Calcolo fondo
    CT_x = []   # Cresta Trasversale
    FT_x = []   # Fondo Trasversale
    AV_x = []   # Area vite
    EV_x = []   # Volume vuoto al giro
    LF_x = []   # Luce fianchi
    LT_x = []   # Luce tetraedrica
    QM_x = []   # Portata massima
    TP_x = []   # Tempo di permanenza zone piene
    TNP_x = []  # Tempo di permanenza zone NON piene
    PL_x = []   # Peso calandrabile
    PL2_x = []  # Peso calandrabile2
    PC_x = []   # Peso di una camera a C
    GT_x = []   # Massimo gradiente velocità di taglio
    CC_x  = []  # Superficie di contatto cilindro in una zona
    CV_x = []   # Superficie di contatto con le viti in una zona
    CV2_x = []  # Superficie di contatto con le viti in una zona (versione 2)
    AEb_x = []  # Angolo elica base vite#
    AEm_x = []  # Angolo elica metà vite#
    AEc_x = []  # Angolo elica cresta vite#
    CPc_x = []  # Canale perpendicolare cresta vite#
    CPb_x = []  # Canale perpendicolare base vite#
    CPm_x = []  # Canale perpendicolare metà vite#
    CRP_x = []  # Cresta perpendicolare#
    CS_x = []   # Calandrate statistiche 
    SR_x = []   # Shear rate 
    Pth_x = []  # Potenza teorica del motore
    Ath_x = []  # Amperes teorici del motore
    nC_x = []    # Numero creste
    Lper_x = []  # Lunghezza
    ASR_x = []   # Area sezione retta cilindro
    

    for l in range(len(DV[k])):                                                                                 #zone sweeping
        CL_x.append(CR[k] * np.sqrt(np.pi**2*DV[k][l]**2+PA[k]**2)/(np.pi*DV[k][l]))                                # [mm ]     Cresta longitudinale 
        FL_x.append(FD[k] * np.sqrt(np.pi**2*(DV[k][l]-2*PF[k][l])**2+PA[k]**2)/(np.pi*(DV[k][l]-2*PF[k][l])))      # [mm]      Fondo longitudinale                                         
        LC_x.append(IA[k][l]-DV[k][l]+PF[k][l])                                                                     # [mm]      Luce di calandra                                         
        AF_x.append((np.arctan((FL_x[l]-CL_x[l])/(2*PF[k][l])))*(180/np.pi))                                                      # [rad]     Angolo di fresa                                         
        TL_x.append((PA[k]/2*NP[k]) - CL_x[l] - ((DV[k][l]-IA[k][l])*(FL_x[l]-CL_x[l])/(2*PF[k][l])))               # [mm]      Traferro longitudinale                                         
        PFc_x.append(LC_x[l]+DV[k][l]-IA[k][l])                                                                     # [mm]      Profondità filetto                                         
        CX_x.append(IA[k][l] / DV[k][l])                                                                            # [-]       Rapporto di compenetrazione X                                         
        CO_x.append((-np.arctan(CX_x[l]/np.sqrt(-CX_x[l]**2+1))+1.5708)*360/np.pi)                                  # [deg]     Angolo compenetrazione                                         
        CP_x.append((DV[k][l]-IA[k][l])*100/DV[k][l])                                                               # [%]       Compenetrazione viti %                                         
        AC_x.append((np.pi*DV[k][l]**2*(0.5-CO_x[l]/720)+IA[k][l]*np.sqrt(DV[k][l]**2-IA[k][l]**2)/2)/100)          # [cm2]     Area cilindro                                          
        cCR_x.append(CR[k]*np.pi*DV[k][l]/(np.sqrt(np.pi**2*DV[k][l]**2+PA[k]**2)))                                 # [mm]      Calcolo cresta                                         
        cFD_x.append(FD[k]*np.pi*(DV[k][l]-2*PF[k][l])/np.sqrt(np.pi**2*(DV[k][l]-2*PF[k][l])**2+PA[k]**2))         # [mm]      Calcolo fondo                                         
        CT_x.append(CR[k]*np.sqrt(np.pi**2*DV[k][l]**2+PA[k]**2)/PA[k])                                             # [mm]      Cresta Trasversale                                         
        FT_x.append(FD[k]*np.sqrt(np.pi**2*(DV[k][l]-2*PF[k][l])**2+PA[k]**2)/PA[k])                                # [mm]      Fondo Trasversale                                         
        AV_x.append(((np.pi*(DV[k][l]-2*PF[k][l])**2/4)+NP[k]*PF[k][l]*(CT_x[l]+FT_x[l])/2)/100)                    # [cm2]     Area vite                                         
        EV_x.append(((AC_x[l]*100)-2*(AV_x[l]*100))*PA[k]/1000)                                                     # [cm3]     Volume vuoto al giro                                         
        LF_x.append(CR[k]*TL_x[l]*np.cos(AF_x[l]*(np.pi/180))/CL_x[l])                                                          # [mm]      Luce fianchi                                         
        LT_x.append((PA[k]/(2*NP[k]))-CL_x[l])                                                                      # [mm]      Luce tetraedrica                                         
        QM_x.append((EV_x[l]*DA*(VR/60))/1000*3600)                                                                 # [kg/h]    Portata massima                                         
        TP_x.append(EV_x[l]*LZ[k]*DA/(QM_x[l]*PA[k]))                                                               # [s]       Tempo di permanenza zone piene                                         
        TNP_x.append(LZ[k]/((VR/60)*PA[k]))                                                                         # [s]       Tempo di permanenza zone NON piene                                         
        PL_x.append((np.sqrt(np.pi**2*(DV[k][l]-2*PF[k][l])**2+PA[k]**2)+np.sqrt(np.pi**2*DV[k][l]**2+PA[k]**2)))   # [kg]      Peso calandrabile                                         
        PL2_x.append(PL_x[l]*LC_x[l]*LZ[k]*CL_x[l]*DE/(2*PA[k])/1000)                                               # [kg]      Peso calandrabile2                                         
        PC_x.append(EV_x[l]*0.6/(2*NP[k]))                                                                          # [kg]      Peso di una camera a C 
        CV_x.append(((FL_x[l]*np.pi*(DV[k][l]-2*PF[k][l]))+(2*np.pi*PF[k][l]*(DV[k][l]-PF[k][l])/np.cos(AF_x[l]*(np.pi/180)))+(CL_x[l]*DV[k][l]*np.pi*CO_x[l]/360)))  # [?]       Superficie di contatto con le viti in una zona 
        CV2_x.append(CV_x[l]*2*LZ[k]*NP[k]/PA[k]/100)
        GT_x.append(np.pi*(VR/60)*PF[k][l]*2/LC_x[l])                                                               # [1/s]     Massimo gradiente velocità di taglio
        CC_x.append(DV[k][l]*np.pi*(360-CO_x[l])*LZ[k]*NP[k]*((PA[k]/NP[k])-CL_x[l])/(PA[k]*180)/100)               # [mm2]     Superficie di contatto cilindro in una zona
        AEb_x.append(np.arctan((PA[k]/NP[k])/(np.pi*(DV[k][l]-(2*PF[k][l]))))*(180/np.pi))                          # [deg]     Angolo elica base vite
        AEm_x.append(np.arctan((PA[k]/NP[k])/(np.pi*(DV[k][l]-(PF[k][l]))))*(180/np.pi))                            # [deg]     Angolo elica metà vite
        AEc_x.append(np.arctan((PA[k]/NP[k])/(np.pi*DV[k][l]))*(180/np.pi))                                         # [deg]     Angolo elica cresta vite
        CPc_x.append(((PA[k]/NP[k])-CR[k])*np.cos(AEc_x[l]*np.pi/180))                                              # [mm]      Canale perpendicolare cresta vite
        CPb_x.append(((PA[k]/NP[k])-CR[k])*np.cos(AEb_x[l]*np.pi/180))                                              # [mm]      Canale perpendicolare base vite
        CPm_x.append(((PA[k]/NP[k])-CR[k])*np.cos(AEm_x[l]*np.pi/180))                                              # [mm]      Canale perpendicolare metà vite
        CRP_x.append(((PA[k]/NP[k]))*np.cos(AEc_x[l]*np.pi/180)-CPc_x[l])                                           # [mm]      Cresta perpendicolare
        CS_x.append(PL_x[l]/PC_x[l])                                                                                # [n°]      Calandrate statistiche
        SR_x.append(np.pi*(VR/60)*PF[k][l]*2/LC_x[l])                                                               # [s^-1]    Shear rate
        Pth_x.append((QM_x[l]/3.6)*(0.41/DE))                                                                       # [kW]      Potenza teorica motore
        Ath_x.append((QM_x[l]/3.6)*(0.81/DE))                                                                       # [A]       Amperes teorici motore
        nC_x.append(LZ[k]/PA[k]*NP[k])                                                                              # [n]       Numero creste
        Lper_x.append(LZ[k]/LS*100)                                                                                 # [%]       Lunghezza percentuale
        ASR_x.append(AC_x[l]/100)                                                                                   # [cm^2]    Area seziione retta cilindro

    CL.append(CL_x)     # CL global list each element is a sublist. CL_x is a list but each element is just singular element. CL_x stores sweeping results inside zone. CL elements indicate zones.
    FL.append(FL_x)
    LC.append(LC_x)
    AF.append(AF_x)
    TL.append(TL_x)
    PFc.append(PFc_x)
    CX.append(CX_x)
    CO.append(CO_x)
    CP.append(CP_x)
    AC.append(AC_x)
    cCR.append(cCR_x)
    cFD.append(cFD_x)
    CT.append(CT_x)
    FT.append(FT_x)
    AV.append(AV_x)
    EV.append(EV_x)
    LF.append(LF_x)
    LT.append(LT_x)
    QM.append(QM_x)
    TP.append(TP_x)
    TNP.append(TNP_x)
    PL.append(PL_x)
    PL2.append(PL2_x)
    PC.append(PC_x)
    GT.append(GT_x)
    CC.append(CC_x)
    CV.append(CV_x)
    CV2.append(CV2_x)
    AEb.append(AEb_x)
    AEm.append(AEm_x)
    AEc.append(AEc_x)
    CPc.append(CPc_x)
    CPb.append(CPb_x)
    CPm.append(CPm_x)
    CRP.append(CRP_x)
    CS.append(CS_x)
    SR.append(SR_x)
    Pth.append(Pth_x)
    Ath.append(Ath_x)
    nC.append(nC_x)
    Lper.append(Lper_x) 
    ASR.append(ASR_x)

# for loop to write cells in excel

#output = [CL, FL, LC, AF, TL, PFc, CX, CO, CP, AC, cCR, cFD, CT, FT, AV, EV, LF, LT, QM, TP, TNP, PL, PL2, PC, GT, CC, CV, CV2]

# Create excel file

from openpyxl import load_workbook # type: ignore
from openpyxl.worksheet.table import Table, TableStyleInfo # type: ignore
from openpyxl.utils import get_column_letter # type: ignore


# Carica il file Excel esistente
workbook = load_workbook(filename, read_only=False, keep_vba=True)
ws = workbook.active    
for z in range(NZ):
    # Crea un nuovo foglio
    new_sheet = workbook.create_sheet(f"ZONA{z+1}")
    data = [
        ["Cresta longitudinale                                       ", "[mm]     ", "CL   "] + CL[z],
        ["Fondo longitudinale                                        ", "[mm]     ", "FL   "] + FL[z],
        ["Luce di calandra                                           ", "[mm]     ", "LC   "] + LC[z],
        ["Angolo di fresa                                            ", "[-]      ", "AF   "] + AF[z],
        ["Traferro longitudinale                                     ", "[mm]     ", "TL   "] + TL[z],
        ["Profondità filetto calcolata                               ", "[mm]     ", "PFc  "] + PFc[z],
        ["Rapporto di compenetrazione X                              ", "[-]      ", "CX   "] + CX[z],
        ["Angolo compenetrazione                                     ", "[deg]    ", "CO   "] + CO[z],
        ["Compenetrazione viti %                                     ", "[-]      ", "CP   "] + CP[z],
        ["Area cilindro                                              ", "[mm^2]   ", "AC   "] + AC[z],
        ["Calcolo cresta                                             ", "[mm]     ", "cCR  "] + cCR[z],
        ["Calcolo fondo                                              ", "[mm]     ", "cFD  "] + cFD[z],
        ["Cresta Trasversale                                         ", "[mm]     ", "CT   "] + CT[z],
        ["Fondo Trasversale                                          ", "[mm]     ", "FT   "] + FT[z],
        ["Area vite                                                  ", "[mm^2]   ", "AV   "] + AV[z],
        ["Volume vuoto al giro                                       ", "[mm^3]   ", "EV   "] + EV[z],
        ["Luce fianchi                                               ", "[mm]     ", "LF   "] + LF[z],
        ["Luce tetraedrica                                           ", "[mm]     ", "LT   "] + LT[z],
        ["Portata massima                                            ", "[kg/h]   ", "QM   "] + QM[z],
        ["Tempo di permanenza zone piene                             ", "[s]      ", "TP   "] + TP[z],
        ["Tempo di permanenza zone NON piene                         ", "[s]      ", "TNP  "] + TNP[z],
        ["Peso calandrabile                                          ", "[kg]     ", "PL   "] + PL[z],
        ["Peso calandrabile2                                         ", "[kg]     ", "PL2  "] + PL2[z],
        ["Peso di una camera a C                                     ", "[kg]     ", "PC   "] + PC[z],
        ["Massimo gradiente velocità di taglio                       ", "[mm/s^2] ", "GT   "] + GT[z],
        ["Superficie di contatto cilindro in una zona                ", "[mm^2]   ", "CC   "] + CC[z],
        ["Superficie di contatto con le viti in una zona             ", "[mm^2]   ", "CV   "] + CV[z],
        ["Superficie di contatto con le viti in una zona (versione 2)", "[mm^2]   ", "CV2  "] + CV2[z],
        ["Angolo elica base vite                                     ", "[deg]    ", "AEb  "] + AEb[z],
        ["Angolo elica metà vite                                     ", "[deg]    ", "AEm  "] + AEm[z],
        ["Angolo elica cresta vite                                   ", "[deg]    ", "AEc  "] + AEc[z],
        ["Canale perpendicolare cresta vite                          ", "[mm]     ", "CPc  "] + CPc[z],
        ["Canale perpendicolare base vite                            ", "[mm]     ", "CPb  "] + CPb[z],
        ["Canale perpendicolare metà vite                            ", "[mm]     ", "CPm  "] + CPm[z],
        ["Cresta perpendicolare                                      ", "[mm]     ", "CRP  "] + CRP[z],
        ["Calandrate statistiche                                     ", "[n°]     ", "CS   "] + CS[z],
        ["Shear rate                                                 ", "[s-1]    ", "SR   "] + SR[z],
        ["Potenza teorica del motore                                 ", "[kW]     ", "Pth  "] + Pth[z],
        ["Amperes teorici del motore                                 ", "[A]      ", "Ath  "] + Ath[z],
        ["Numero creste                                              ", "[n]      ", "nC   "] + nC[z],
        ["Lunghezza                                                  ", "[%]      ", "Lper "] + Lper[z],
        ["Area sezione retta cilindro                                ", "[cm^2]   ", "ASR  "] + ASR[z]

    ]
    
    # Calcola il numero di elementi in CL
    num_elements = len(CL[z])
    # Crea i dati per gli header dinamici
    dynamic_headers = [(i+1) for i in range(num_elements)]
    # Aggiungi il tuo header dinamico insieme agli altri header

    new_sheet.append(["", "", ""] + dynamic_headers)
    new_sheet.append(["OUTPUT", "Units", "Code"])

    for row in data:
        new_sheet.append(row)

    # Calcola la lettera della colonna corrispondente all'indice massimo
    num_columns = num_elements // 26
    if num_elements % 26 != 0:
        num_columns += 1
    last_column_letter = get_column_letter(num_elements + 3)

    #tab = Table(displayName= f"ZONA{z+1}", ref=f"D1:{last_column_letter}29")

    #new_sheet.add_table(tab)
workbook.save(filename[0:-5]+'_OUTPUT.xlsm')

print("\nOUTPUT successfully computed\n")
time.sleep(0.5)

